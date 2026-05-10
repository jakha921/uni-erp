"""Reusable Excel export/import utilities."""

import io
from datetime import date, datetime
from typing import Any

import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

HEADER_FILL = PatternFill(start_color="2DB976", end_color="2DB976", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def export_to_excel(
    data: list[dict[str, Any]],
    columns: list[tuple[str, str]],
    filename: str = "export",
    sheet_name: str = "Ma'lumotlar",
) -> HttpResponse:
    """Export list of dicts to Excel file as HttpResponse.

    Args:
        data: List of dicts with data
        columns: List of (key, header_label) tuples
        filename: Output filename without extension
        sheet_name: Excel sheet name
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Headers
    for col_idx, (_, label) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # Data rows
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, (key, _) in enumerate(columns, 1):
            value = row_data.get(key, "")
            if isinstance(value, (date, datetime)):
                value = value.strftime("%d.%m.%Y")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")

    # Auto-width columns
    for col_idx, (_, label) in enumerate(columns, 1):
        max_len = len(label)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 3, 50)

    # Freeze header row
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
    return response


def parse_excel(file, columns: list[tuple[str, str]]) -> list[dict[str, Any]]:
    """Parse uploaded Excel file into list of dicts.

    Args:
        file: Uploaded file object
        columns: List of (key, header_label) tuples — maps header to dict key

    Returns:
        List of dicts with parsed data
    """
    wb = openpyxl.load_workbook(file, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip() if h else "" for h in rows[0]]
    label_to_key = {label: key for key, label in columns}

    result = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        item = {}
        for col_idx, value in enumerate(row):
            if col_idx < len(headers):
                header = headers[col_idx]
                key = label_to_key.get(header)
                if key:
                    item[key] = value
        if item:
            result.append(item)

    return result
